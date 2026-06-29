# ============================================================
# STM Image to Excel Parser — Streamlit Edition
# KBANK + KRUNGSRI + BBL + SCB
# รัน: streamlit run app.py
# ============================================================

import os
import re
import json
import math
import tempfile
import io
import base64
import streamlit as st
import pandas as pd
from datetime import datetime
from PIL import Image, ImageEnhance, ImageFilter, ImageOps

# ─── Page config ────────────────────────────────────────────
st.set_page_config(
    page_title="STM Parser",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ─── CSS ────────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        max-width: 980px;
    }

    h1 {
        font-size: 2.35rem !important;
        font-weight: 800 !important;
        letter-spacing: -0.5px;
        margin-bottom: 0.25rem;
    }

    div[data-testid="stCaptionContainer"] {
        color: var(--text-color-secondary, #8c8c8c);
        font-size: 0.95rem;
    }

    .section-title {
        display: flex;
        align-items: center;
        gap: 8px;
        font-weight: 800;
        font-size: 1.15rem;
        line-height: 1.4;
        margin: 18px 0 10px 0;
        color: inherit !important;
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
    }

    .section-title .step-number {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        min-width: 30px;
        height: 30px;
        border-radius: 999px;
        background: #4c82fb;
        color: #ffffff;
        font-size: 0.95rem;
        font-weight: 800;
    }

    .section-title .step-text { color: inherit; }

    .badge {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        background: #e8f0fe;
        color: #1a56db;
        border-radius: 999px;
        padding: 7px 18px;
        font-size: 1.05rem;
        font-weight: 900;
        line-height: 1.2;
        margin-left: 10px;
        min-width: 68px;
        min-height: 34px;
        letter-spacing: 0.2px;
    }

    @media (prefers-color-scheme: dark) {
        .badge {
            background: rgba(76, 130, 251, 0.18);
            color: #a9c4ff;
            border: 1px solid rgba(169, 196, 255, 0.25);
        }
    }

    div[data-testid="stRadio"] label {
        font-size: 1.02rem !important;
        font-weight: 600 !important;
        padding: 4px 8px;
        border-radius: 8px;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] { gap: 14px; }

    div[data-testid="stFileUploader"] { margin-top: 4px; }
    div[data-testid="stFileUploader"] section { border-radius: 10px; }

    div[data-testid="stAlert"] { border-radius: 10px; }

    div[data-testid="stButton"] button,
    div[data-testid="stDownloadButton"] button {
        border-radius: 10px;
        font-weight: 750;
    }

    div[data-testid="metric-container"] {
        border-radius: 10px;
        padding: 14px 16px !important;
    }

    div[data-testid="stDataFrame"] {
        border-radius: 10px;
        overflow: hidden;
    }

    .muted-note {
        color: var(--text-color-secondary, #8c8c8c);
        font-size: 0.9rem;
        line-height: 1.6;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# 0)  Constants & Helpers
# ============================================================
FORCE_YEAR          = datetime.now().year
BALANCE_TOLERANCE   = 0.05
LOW_CONF_THRESHOLD  = 0.75

BANK_CONFIGS = {
    "DEFAULT": {
        "bank_keywords": [],
        "opening_keywords": ["ยอด ยก มา","ยอดยกมา","ยอด ยก","ยก มา","BALANCE BROUGHT FORWARD","BROUGHT FORWARD","BALANCE B/F","B/F"],
        "credit_keywords": ["รับ โอน","รับโอน","รับ โอน เงิน","ฝาก","ฝาก เงิน","ฝาก เงินสด","ฝากเงินสด","ดอกเบี้ย","คืนเงิน","เงิน เข้า","เงินเข้า","CREDIT","DEPOSIT","TRANSFER IN"],
        "debit_keywords":  ["โอน เงิน","โอนเงิน","โอน เงิน พร้อม เพ ย์","โอน เงิน พร้อมเพย์","จ่าย บิล","จ่าย คิว อา ร์","จ่าย QR","ชำระ","ชา ระ","ชํา ระ","ถอน","ถอนเงิน","หัก","ค่าธรรมเนียม","จ่าย","DEBIT","WITHDRAW","WITHDRAWAL","PAYMENT","FEE","TRANSFER OUT"],
    },
    "KBANK": {
        "bank_keywords": ["KASIKORN","KBANK","K PLUS","กสิกร","ธนาคาร กสิกร ไทย","ธนาคารกสิกรไทย"],
        "opening_keywords": ["ยอด ยก มา","ยอดยกมา","ยอด ยก","ยก มา","BALANCE BROUGHT FORWARD","BROUGHT FORWARD"],
        "credit_keywords": ["รับ โอน","รับโอน","ฝาก","ฝาก เงินสด","เงิน เข้า","เงินเข้า","ดอกเบี้ย","CREDIT","DEPOSIT"],
        "debit_keywords":  ["โอน เงิน","โอนเงิน","โอน เงิน พร้อม เพ ย์","โอน เงิน พร้อมเพย์","จ่าย คิว อา ร์","จ่าย QR","ถอน","ถอนเงิน","หัก","ชำระ","ค่าธรรมเนียม","DEBIT","WITHDRAW","PAYMENT","FEE"],
    },
    "KRUNGSRI": {
        "bank_keywords": ["KRUNGSRI","กรุงศรี","กรุง ศรี","BANK OF AYUDHYA","AYUDHYA","MOBILE KRUNGSRI","KRUNGSRI VISA","FIRSTCHOICE","FIRST CHOICE"],
        "opening_keywords": ["ยอด ยก มา","ยอดยกมา","ยอด ยก","ยก มา","BALANCE BROUGHT FORWARD","BROUGHT FORWARD","BALANCE B/F","B/F"],
        "credit_keywords": ["รับ โอน เงิน","รับ โอน","รับโอน","ฝาก เงิน","ฝาก","ดอกเบี้ย เงิน ฝาก","ดอกเบี้ย","คืนเงิน","เงิน เข้า","เงินเข้า","CREDIT","DEPOSIT","TRANSFER IN"],
        "debit_keywords":  ["จ่าย บิล","จ่าย คิว อา ร์","จ่าย QR","โอน เงิน พร้อม เพ ย์","โอน เงิน พร้อมเพย์","โอน เงิน","โอนเงิน","ถอนเงิน","ถอน","หัก","ค่าธรรมเนียม","ชำระ","จ่าย","DEBIT","WITHDRAW","WITHDRAWAL","PAYMENT","FEE"],
    },
    "BBL": {
        "bank_keywords": ["BANGKOK BANK","BANGKOKBANK","ธนาคารกรุงเทพ","ธนาคาร กรุงเทพ","บัวหลวง","BUALUANG","STATEMENT OF SAVING ACCOUNT"],
        "opening_keywords": ["B/F","B / F","BALANCE B/F","BALANCE BROUGHT FORWARD","BROUGHT FORWARD","ยอด ยก มา","ยอดยกมา","ยอด ยก","ยก มา"],
        "credit_keywords": ["TRF FR OTH BK","TRF FROTH BK","TRF FROM OTH BK","TRF FROM OTHER BANK","TRANSFER IN","SALARY","SMART","CREDIT","DEPOSIT","ฝาก","รับ โอน","รับโอน","เงิน เข้า","เงินเข้า"],
        "debit_keywords":  ["TRF TO OTH BK","TRF TOOTH BK","TRF TO OTHER BANK","TRF. PROMPTPAY","TRF . PROMPTPAY","PMT. PROMPTPAY","PMT . PROMPTPAY","PMT.PROMPTPAY","PMT FOR GOODS","CASH W/D ATM","CASH W / D ATM","WITHDRAWAL","WITHDRAW","TRANSFER","PAYMENT","DEBIT","FEE","ถอน","ถอนเงิน","โอน เงิน","โอนเงิน","ชำระ","จ่าย"],
    },
    "SCB": {
        "bank_keywords": ["SCB","ไทยพาณิชย์","ไทย พาณิชย์","ธนาคารไทยพาณิชย์","ธนาคาร ไทย พาณิชย์","THE SIAM COMMERCIAL BANK","SIAM COMMERCIAL BANK"],
        "opening_keywords": ["ยอดเงินคงเหลือยกมา","ยอด เงิน คงเหลือ ยก มา","BALANCE BROUGHT FORWARD","BROUGHT FORWARD","BALANCE B/F","B/F"],
        "credit_keywords": [" X1 ","X1","รับ โอน จาก","รับโอนจาก","รับ โอน","รับโอน","โอน จาก","โอนจาก","ฝาก","เงิน เข้า","เงินเข้า","CREDIT","DEPOSIT","TRANSFER IN"],
        "debit_keywords":  [" X2 ","X2","โอน ไป","โอนไป","PromptPay","PROMPTPAY","จ่าย บิล","จ่ายบิล","จ่าย","ถอน","ถอนเงิน","ชำระ","ชา ระ","ชํา ระ","PAYMENT","WITHDRAW","WITHDRAWAL","TRANSFER OUT","FEE"],
    },
}

NOISE_KEYWORDS = [
    "STATEMENT OF SAVING ACCOUNT","THE SIAM COMMERCIAL BANK","ACCOUNT NO","เลขที่บัญชี",
    "ชื่อ - สกุล","ADDRESS","สาขา","BRANCH","DATE | TIME","DEBIT/CREDIT","BALANCE/BAHT",
    "DESCRIPTION/NOTE","TOTAL AMOUNTS","TOTAL ITEMS","THIS DOCUMENT IS AUTO-GENERATED",
    "เอกสาร ฉบับ นี้","ออก โดย ระบบ อัตโนมัติ",
]

BANK_LABELS = {
    "AUTO":    "ตรวจจับอัตโนมัติ",
    "KBANK":   "กสิกรไทย (KBANK)",
    "KRUNGSRI":"กรุงศรี (Krungsri)",
    "BBL":     "กรุงเทพ (BBL)",
    "SCB":     "ไทยพาณิชย์ (SCB)",
}


# ============================================================
# 1)  Basic Helpers
# ============================================================
def clean_text(s: str) -> str:
    s = str(s).replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()

def get_config(bank: str) -> dict:
    return BANK_CONFIGS.get(bank, BANK_CONFIGS["DEFAULT"])

def normalize_amount_text(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"\s+", "", s)
    ocr_map = {"O":"0","o":"0","I":"1","l":"1","|":"1","S":"5","s":"5","B":"8","g":"9","q":"9","Z":"2","z":"2"}
    for c, d in ocr_map.items():
        s = s.replace(c, d)
    if re.match(r"^\d{1,3}(,\d{3})+,\d{2}$", s):
        s = s[:-3] + "." + s[-2:]
    if re.match(r"^\d+,\d{2}$", s):
        s = s.replace(",", ".")
    if re.match(r"^\d{1,3}\.\d{3}\.\d{2}$", s):
        parts = s.split(".")
        s = parts[0] + "," + parts[1] + "." + parts[2]
    return s

def parse_money(s) -> float | None:
    if s is None or (isinstance(s, float) and math.isnan(s)):
        return None
    s = normalize_amount_text(str(s)).replace(" ", "")
    s = re.sub(r"[^0-9,.\-]", "", s)
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except:
        return None

def is_money_token(s: str) -> bool:
    s = normalize_amount_text(str(s).strip())
    pats = [
        r"^-?\d{1,3}(,\d{3})*(\.\d{2})$",
        r"^-?\d+\.\d{2}$",
        r"^-?\d+,\d{2}$",
        r"^-?\d{1,3}\.\d{3}\.\d{2}$",
    ]
    return any(re.match(p, s) for p in pats)

def extract_money_loose(token: str) -> list[float]:
    s = str(token)
    ocr_map = {"O":"0","o":"0","I":"1","l":"1","|":"1","S":"5","s":"5","B":"8","g":"9","q":"9","Z":"2","z":"2"}
    for c, d in ocr_map.items():
        s = s.replace(c, d)
    priority_patterns = [
        r"(?<![\d,])-?\d{1,3}(?:,\d{3})+\.\d{2}(?!\d)",
        r"(?<![\d.])-?\d{1,3}(?:\.\d{3})+,\d{2}(?!\d)",
        r"(?<![\d.])-?\d{1,3}(?:\.\d{3})+\.\d{2}(?!\d)",
        r"(?<![\d,])-?\d+\.\d{2}(?!\d)",
        r"(?<![\d,])-?\d+,\d{2}(?![\d.])",
    ]
    for pat in priority_patterns:
        matches = re.findall(pat, s)
        if matches:
            vals = [v for m in matches for v in [parse_money(m)] if v is not None]
            if vals:
                return vals
    return []

def money_to_key(value) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    try:
        return str(int(round(float(value) * 100)))
    except:
        return ""

def is_valid_number(x) -> bool:
    if x is None:
        return False
    try:
        if math.isnan(float(x)):
            return False
        return True
    except:
        return False


# ============================================================
# 2)  Date Helpers
# ============================================================
def fix_ocr_date(s: str) -> str:
    s = str(s).strip()
    s = s.replace("O","0").replace("o","0").replace("I","1").replace("l","1")
    s = s.replace(".","-").replace("/","-")
    return s

def is_date_token(s: str) -> bool:
    s = fix_ocr_date(s)
    return bool(re.match(r"^\d{1,2}[-]\d{1,2}[-]\d{2,4}$", s) or
                re.match(r"^\d{1,2}[-]\d{1,2}$", s))

def normalize_date(s: str) -> str | None:
    s = fix_ocr_date(s)
    m1 = re.match(r"^(\d{1,2})[-](\d{1,2})[-](\d{2,4})$", s)
    m2 = re.match(r"^(\d{1,2})[-](\d{1,2})$", s)
    if m1:
        day, month = int(m1.group(1)), int(m1.group(2))
    elif m2:
        day, month = int(m2.group(1)), int(m2.group(2))
    else:
        return None
    try:
        dt = datetime(FORCE_YEAR, month, day)
    except:
        return None
    return dt.strftime("%d-%m-%Y")

def extract_date(line_text: str) -> str | None:
    for tok in str(line_text).split():
        d = normalize_date(tok)
        if d:
            return d
    return None

def is_time_token(s: str) -> bool:
    return bool(re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", str(s).strip()))

def extract_time(line_text: str) -> str | None:
    for tok in str(line_text).split():
        if is_time_token(tok):
            return tok
    return None

def parse_date_for_sort(date_text: str):
    if not date_text or pd.isna(date_text):
        return pd.NaT
    return pd.to_datetime(str(date_text).strip().replace("/","-"), format="%d-%m-%Y", errors="coerce")


# ============================================================
# 3)  Text / Transaction Helpers
# ============================================================
def has_opening_text(text: str, bank: str = "DEFAULT") -> bool:
    up = clean_text(text).upper()
    for kw in get_config(bank).get("opening_keywords", []):
        if kw.upper() in up:
            return True
    for kw in BANK_CONFIGS["DEFAULT"]["opening_keywords"]:
        if kw.upper() in up:
            return True
    return False

def classify_kw(text: str, bank: str = "DEFAULT") -> str:
    up = clean_text(text).upper()
    for kw in get_config(bank).get("credit_keywords", []):
        if kw.upper() in up: return "credit"
    for kw in get_config(bank).get("debit_keywords", []):
        if kw.upper() in up: return "debit"
    for kw in BANK_CONFIGS["DEFAULT"]["credit_keywords"]:
        if kw.upper() in up: return "credit"
    for kw in BANK_CONFIGS["DEFAULT"]["debit_keywords"]:
        if kw.upper() in up: return "debit"
    return "unknown"

def is_noise_line(text: str) -> bool:
    up = clean_text(text).upper()
    for kw in NOISE_KEYWORDS:
        if kw.upper() in up:
            return True
    if re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s*[-–]\s*\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", up):
        return True
    if re.search(r"หน้า\s*\d+\s*/\s*\d+", up):
        return True
    return False

def detect_bank(full_text: str) -> str:
    text = clean_text(full_text).upper()
    best, best_score = "DEFAULT", 0
    for bank_name, cfg in BANK_CONFIGS.items():
        if bank_name == "DEFAULT":
            continue
        score = 0
        for kw in cfg.get("bank_keywords", []):
            if kw.upper() in text: score += 10
        for kw in cfg.get("debit_keywords", []):
            if kw.upper() in text: score += 1
        for kw in cfg.get("credit_keywords", []):
            if kw.upper() in text: score += 1
        if score > best_score:
            best_score = score
            best = bank_name
    return best


# ============================================================
# 4)  Healing Engine
# ============================================================
def try_healing_amount(prev_balance, read_amount, balance, tolerance=0.05):
    if any(v is None or (isinstance(v, float) and math.isnan(v))
           for v in [prev_balance, read_amount, balance]):
        return None, None
    exp_debit  = round(prev_balance - balance, 2)
    exp_credit = round(balance - prev_balance, 2)
    exp_amount = exp_debit if exp_debit > 0 else exp_credit
    tx_type    = "debit" if exp_debit > 0 else "credit"
    str_read, str_exp = f"{read_amount:.2f}", f"{exp_amount:.2f}"
    if abs(read_amount / 100 - exp_amount) <= tolerance:
        return exp_amount, f"healed_missing_dot_{tx_type}"
    if len(str_read) == len(str_exp):
        mismatches = sum(1 for a, b in zip(str_read, str_exp) if a != b)
        if mismatches <= 1:
            return exp_amount, f"healed_swapped_digit_{tx_type}"
    if len(str_exp) > len(str_read) and (str_read in str_exp or str_exp.endswith(str_read)):
        return exp_amount, f"healed_omitted_digit_{tx_type}"
    return None, None

def suggest_from_balance(prev_balance, balance):
    if any(v is None or (isinstance(v, float) and math.isnan(v)) for v in [prev_balance, balance]):
        return None, None
    diff = round(balance - prev_balance, 2)
    if diff > 0: return abs(diff), "credit"
    if diff < 0: return abs(diff), "debit"
    return 0.0, "zero"

def amount_balance_match(prev_balance, amount, balance, tolerance=BALANCE_TOLERANCE):
    if any(v is None or (isinstance(v, float) and math.isnan(v)) for v in [prev_balance, amount, balance]):
        return None
    debit_exp  = round(prev_balance - amount, 2)
    credit_exp = round(prev_balance + amount, 2)
    if abs(balance - debit_exp) <= tolerance:
        return {"type": "debit",   "expected_balance": debit_exp,  "diff": round(balance - debit_exp, 2),  "balance_check": "OK_DEBIT"}
    if abs(balance - credit_exp) <= tolerance:
        return {"type": "credit",  "expected_balance": credit_exp, "diff": round(balance - credit_exp, 2), "balance_check": "OK_CREDIT"}
    return None


# ============================================================
# 5)  Image Pre-processing
# ============================================================
def preprocess_image_variants(pil_img: Image.Image, base_name: str) -> list[tuple[str, Image.Image]]:
    img = pil_img.convert("RGB")
    variants = [("original", img)]

    i1 = img.resize((img.width*2, img.height*2), Image.LANCZOS)
    i1 = ImageEnhance.Contrast(i1).enhance(1.9)
    i1 = ImageEnhance.Sharpness(i1).enhance(2.2)
    variants.append(("2x_contrast_sharp", i1))

    i2 = img.resize((img.width*3, img.height*3), Image.LANCZOS)
    i2 = i2.convert("L"); i2 = ImageOps.autocontrast(i2)
    i2 = ImageEnhance.Contrast(i2).enhance(1.8)
    i2 = ImageEnhance.Sharpness(i2).enhance(2.0)
    variants.append(("3x_gray_autocontrast", i2))

    i3 = img.resize((img.width*3, img.height*3), Image.LANCZOS)
    i3 = i3.convert("L"); i3 = ImageOps.autocontrast(i3)
    i3 = i3.filter(ImageFilter.UnsharpMask(radius=1.2, percent=180, threshold=3))
    variants.append(("3x_unsharp", i3))

    i4 = img.resize((img.width*3, img.height*3), Image.LANCZOS)
    i4 = i4.convert("L"); i4 = ImageOps.autocontrast(i4)
    i4 = ImageEnhance.Contrast(i4).enhance(2.0)
    i4 = i4.point(lambda x: 255 if x > 175 else 0)
    variants.append(("3x_bw_175", i4))

    i5 = img.resize((img.width*3, img.height*3), Image.LANCZOS)
    i5 = i5.convert("L"); i5 = ImageOps.autocontrast(i5)
    i5 = ImageEnhance.Contrast(i5).enhance(1.7)
    i5 = i5.point(lambda x: 255 if x > 145 else 0)
    variants.append(("3x_bw_145", i5))

    return variants


# ============================================================
# 6)  Google Vision OCR
# ============================================================
def get_vision_client():
    from google.cloud import vision
    from google.oauth2 import service_account

    if "gcp_service_account" not in st.secrets:
        st.error("ไม่พบ [gcp_service_account] ใน Streamlit Secrets")
        st.stop()

    try:
        info = dict(st.secrets["gcp_service_account"])
        if "private_key" in info and isinstance(info["private_key"], str):
            info["private_key"] = info["private_key"].replace("\\n", "\n")
        creds = service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
    except Exception as e:
        st.error(f"อ่าน [gcp_service_account] ไม่สำเร็จ: {e}")
        st.stop()

    return vision.ImageAnnotatorClient(credentials=creds)


def vision_ocr_pil(pil_img: Image.Image, client, file_name: str) -> tuple[str, pd.DataFrame]:
    from google.cloud import vision as vis
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    content = buf.getvalue()

    image = vis.Image(content=content)
    response = client.document_text_detection(
        image=image,
        image_context=vis.ImageContext(language_hints=["th", "en"]),
    )
    if response.error.message:
        raise RuntimeError(response.error.message)

    rows = []
    for page in response.full_text_annotation.pages:
        for block in page.blocks:
            for para in block.paragraphs:
                for word in para.words:
                    word_text = "".join(s.text for s in word.symbols)
                    verts = word.bounding_box.vertices
                    xs = [v.x for v in verts]; ys = [v.y for v in verts]
                    xmin,xmax = min(xs),max(xs); ymin,ymax = min(ys),max(ys)
                    rows.append({
                        "file_name": file_name,
                        "page_width": page.width, "page_height": page.height,
                        "text": word_text,
                        "word_confidence": word.confidence,
                        "x_min": xmin, "y_min": ymin, "x_max": xmax, "y_max": ymax,
                        "x_center": (xmin+xmax)/2, "y_center": (ymin+ymax)/2,
                        "width": xmax-xmin, "height": ymax-ymin,
                    })
    return response.full_text_annotation.text, pd.DataFrame(rows)


# ============================================================
# 7)  Line Rebuild + Parse
# ============================================================
def rebuild_lines_by_y(words_df: pd.DataFrame, y_threshold: int = 10) -> pd.DataFrame:
    all_lines = []
    for file_name, fg in words_df.groupby("file_name"):
        if fg.empty: continue
        page_width = fg["page_width"].dropna().iloc[0]
        fg = fg[fg["x_center"] >= page_width * 0.04].copy()
        if fg.empty: continue
        fg = fg.sort_values(["y_center","x_center"]).reset_index(drop=True)
        current_line, current_y, line_id = [], None, 0
        for _, row in fg.iterrows():
            y = row["y_center"]
            if current_y is None:
                current_line, current_y = [row], y; continue
            if abs(y - current_y) <= y_threshold:
                current_line.append(row)
                current_y = (current_y + y) / 2
            else:
                ldf = pd.DataFrame(current_line); ldf["line_id"] = line_id
                all_lines.append(ldf); line_id += 1
                current_line, current_y = [row], y
        if current_line:
            ldf = pd.DataFrame(current_line); ldf["line_id"] = line_id
            all_lines.append(ldf)
    return pd.concat(all_lines, ignore_index=True) if all_lines else pd.DataFrame()


def parse_lines(words_df: pd.DataFrame, active_bank: str) -> pd.DataFrame:
    line_words = rebuild_lines_by_y(words_df, y_threshold=10)
    if line_words.empty: return pd.DataFrame()
    line_rows = []
    for (file_name, line_id), g in line_words.groupby(["file_name","line_id"]):
        g = g.sort_values("x_center").copy()
        raw = " ".join(g["text"].astype(str).tolist()).strip()
        raw = re.sub(r"\s+", " ", raw)
        if is_noise_line(raw): continue
        date = extract_date(raw)
        time = extract_time(raw)
        is_opening = has_opening_text(raw, active_bank)
        if not date and not is_opening: continue
        money_vals = []
        for _, row in g.sort_values("x_center").iterrows():
            txt = normalize_amount_text(str(row["text"]))
            if is_money_token(txt) or extract_money_loose(txt):
                v = parse_money(txt)
                if v is None:
                    vs = extract_money_loose(txt)
                    if vs: v = vs[0]
                if v is not None:
                    money_vals.append(v)
        amount, balance = None, None
        if is_opening:
            if money_vals: balance = money_vals[-1]
        else:
            if len(money_vals) >= 2: amount, balance = money_vals[0], money_vals[-1]
            elif len(money_vals) == 1: amount = money_vals[0]
        avg_conf = float(g["word_confidence"].mean()) if "word_confidence" in g.columns else None
        min_conf = float(g["word_confidence"].min()) if "word_confidence" in g.columns else None
        line_rows.append({
            "file_name": file_name, "line_id": int(line_id),
            "date": date, "time": time, "code": None,
            "amount": amount, "balance": balance,
            "money_tokens": " | ".join(str(v) for v in money_vals),
            "money_count": len(money_vals),
            "raw_line_text": raw,
            "y_center": float(g["y_center"].mean()),
            "parsed_date": parse_date_for_sort(date),
            "is_opening_balance": is_opening,
            "line_confidence": avg_conf,
            "min_confidence": min_conf,
        })
    if not line_rows: return pd.DataFrame()
    df = pd.DataFrame(line_rows)
    return df.sort_values(["file_name","y_center"], na_position="last").reset_index(drop=True)


# ============================================================
# 8)  Dedup + Chain Check
# ============================================================
def build_duplicate_key(row) -> str:
    time_key    = str(row.get("time", "") or "").strip()
    amount_key  = money_to_key(row.get("amount"))
    balance_key = money_to_key(row.get("balance"))
    raw         = str(row.get("raw_line_text", ""))
    if has_opening_text(raw):
        return "OPENING|" + balance_key
    if time_key:
        return f"TX|{time_key}|{amount_key}|{balance_key}"
    return f"TX|{amount_key}|{balance_key}"

def deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    df = df.copy()
    df["_dup_key"] = df.apply(build_duplicate_key, axis=1)
    df = df.drop_duplicates(subset=["_dup_key"], keep="first").reset_index(drop=True)
    return df.drop(columns=["_dup_key"], errors="ignore")


def is_opening_check_row(row, active_bank: str = "DEFAULT") -> bool:
    raw_line_text = str(row.get("raw_line_text", ""))
    code = str(row.get("code", ""))
    balance_check = str(row.get("balance_check", ""))
    return (
        has_opening_text(raw_line_text, active_bank)
        or code == "OPENING"
        or balance_check == "OPENING_BALANCE"
    )


def build_final_parsed_stm_from_check(check_df: pd.DataFrame, active_bank: str = "DEFAULT") -> pd.DataFrame:
    """
    ใช้ logic จากโค้ดที่ 1:
    - parsed_stm ต้องสร้างจาก check_df ที่ผ่านการเรียงและตรวจ balance-chain แล้ว
    - คอลัมน์ปลายทางต้องเหลือแค่ date, debit, credit, balance
    - ห้าม sort วัน/เวลาใหม่เอง ให้รักษาลำดับ seq จาก check_df
    """
    parsed_cols = ["date", "debit", "credit", "balance"]

    if check_df is None or check_df.empty:
        return pd.DataFrame(columns=parsed_cols)

    df = check_df.copy()
    required_cols = [
        "seq", "date", "debit", "credit", "balance",
        "balance_check", "raw_line_text", "code"
    ]
    for col in required_cols:
        if col not in df.columns:
            df[col] = None

    opening_mask = df.apply(lambda r: is_opening_check_row(r, active_bank), axis=1)
    df = df[~opening_mask].copy()

    df = df[df["balance"].apply(is_valid_number)].copy()

    has_debit_or_credit = (
        df["debit"].apply(is_valid_number)
        | df["credit"].apply(is_valid_number)
    )
    df = df[has_debit_or_credit].copy()

    df = df.sort_values("seq").reset_index(drop=True)
    final_df = df[parsed_cols].copy()

    for col in ["debit", "credit", "balance"]:
        final_df[col] = pd.to_numeric(final_df[col], errors="coerce")

    return final_df.reset_index(drop=True)


def add_debit_credit_check(ordered_df: pd.DataFrame, active_bank: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    สร้าง check_df ก่อน แล้วค่อยสร้าง parsed_stm จาก check_df อีกที
    เพื่อให้ผลลัพธ์/ชีท/คอลัมน์ตรงกับโค้ดที่ 1
    """
    check_cols = [
        "seq", "date", "time", "code", "debit", "credit", "balance",
        "prev_balance", "amount", "suggested_amount", "suggested_type",
        "expected_balance", "diff", "line_confidence", "min_confidence",
        "money_tokens", "raw_line_text", "order_method", "balance_check"
    ]

    if ordered_df is None or ordered_df.empty:
        return (
            pd.DataFrame(columns=["date", "debit", "credit", "balance"]),
            pd.DataFrame(columns=check_cols),
        )

    check_rows = []
    prev_balance = None
    seq = 0

    for _, row in ordered_df.iterrows():
        seq += 1
        date         = row.get("date")
        time         = row.get("time")
        code         = row.get("code")
        amount       = row.get("amount")
        bal_val      = row.get("balance")
        raw          = row.get("raw_line_text", "")
        money_tokens = row.get("money_tokens", "")
        line_conf    = row.get("line_confidence")
        min_conf     = row.get("min_confidence")
        order_method = row.get("order_method", "")
        is_opening   = row.get("is_opening_balance", False)

        debit = credit = expected_balance = diff_ = suggested_amount = suggested_type = None
        balance_check = ""

        low_conf = (
            (line_conf is not None and not pd.isna(line_conf) and line_conf < LOW_CONF_THRESHOLD) or
            (min_conf  is not None and not pd.isna(min_conf)  and min_conf  < LOW_CONF_THRESHOLD)
        )

        if is_opening or code == "OPENING" or has_opening_text(raw, active_bank):
            balance_check = "OPENING_BALANCE"
        else:
            match = amount_balance_match(prev_balance, amount, bal_val)
            if match:
                expected_balance = match["expected_balance"]
                diff_            = match["diff"]
                balance_check    = match["balance_check"]
                if match["type"] == "debit":
                    debit = amount
                else:
                    credit = amount
            else:
                h_amt, h_type = try_healing_amount(prev_balance, amount, bal_val)
                if h_amt is not None:
                    if "debit" in h_type:
                        debit = h_amt
                    else:
                        credit = h_amt
                    expected_balance = bal_val
                    diff_ = 0.0
                    balance_check = f"OK_{h_type.upper()}"
                else:
                    tx_type = classify_kw(raw, active_bank)
                    s_amt, s_type = suggest_from_balance(prev_balance, bal_val)
                    if s_amt is not None and s_type in ("debit", "credit"):
                        suggested_amount, suggested_type = s_amt, s_type
                        expected_balance = round(
                            (prev_balance - s_amt) if s_type == "debit" else (prev_balance + s_amt),
                            2,
                        ) if prev_balance is not None else None
                        diff_ = round(bal_val - expected_balance, 2) if expected_balance is not None else None

                        if tx_type == s_type:
                            balance_check = "SUGGEST_AMOUNT_BY_KEYWORD_REVIEW"
                            if s_type == "debit":
                                debit = s_amt
                            else:
                                credit = s_amt
                        else:
                            balance_check = "SUGGEST_AMOUNT_REVIEW"
                    else:
                        balance_check = "CHAIN_BROKEN_OCR_REVIEW"

        if low_conf and "OK" in balance_check:
            balance_check += "_LOW_CONFIDENCE"

        check_rows.append({
            "seq": seq,
            "date": date,
            "time": time,
            "code": code,
            "debit": debit,
            "credit": credit,
            "balance": bal_val,
            "prev_balance": prev_balance,
            "amount": amount,
            "suggested_amount": suggested_amount,
            "suggested_type": suggested_type,
            "expected_balance": expected_balance,
            "diff": diff_,
            "line_confidence": line_conf,
            "min_confidence": min_conf,
            "money_tokens": money_tokens,
            "raw_line_text": raw,
            "order_method": order_method,
            "balance_check": balance_check,
        })

        if bal_val is not None and not pd.isna(bal_val):
            prev_balance = bal_val

    check_df = pd.DataFrame(check_rows, columns=check_cols)
    parsed_df = build_final_parsed_stm_from_check(check_df, active_bank)
    return parsed_df, check_df


# ============================================================
# 9)  Score + Best Variant
# ============================================================
def score_words(words_df: pd.DataFrame, active_bank: str = "DEFAULT") -> float:
    if words_df.empty: return -999999.0
    parsed = parse_lines(words_df, active_bank)
    if parsed.empty: return 0.0
    conf = parsed["line_confidence"].fillna(0).mean() * 100 if "line_confidence" in parsed.columns else 0
    return (
        len(parsed) * 100 +
        parsed["money_count"].fillna(0).sum() * 20 +
        parsed["date"].notna().sum() * 50 +
        conf
    )


# ============================================================
# 10)  Full Pipeline
# ============================================================
def run_pipeline(
    uploaded_files,
    selected_bank: str,
    progress_cb=None,
    status_cb=None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    client = get_vision_client()
    all_words_list, all_full_texts = [], []
    n = len(uploaded_files)

    for i, uf in enumerate(uploaded_files):
        fname = uf.name
        if status_cb: status_cb(f"📡 OCR ภาพที่ {i+1}/{n}: {fname}")
        if progress_cb: progress_cb((i) / n * 0.6)

        pil_img = Image.open(uf).convert("RGB")
        variants = preprocess_image_variants(pil_img, fname)

        best_score, best_words = -999999, None

        for v_name, v_img in variants:
            try:
                full_text, word_df = vision_ocr_pil(v_img, client, fname)
                if word_df.empty: continue
                score = score_words(word_df, selected_bank if selected_bank != "AUTO" else "DEFAULT")
                if score > best_score:
                    best_score = score
                    best_words = word_df
                    best_text  = full_text
            except Exception:
                pass

        if best_words is None:
            raise RuntimeError(f"OCR ล้มเหลวทุก variant สำหรับ {fname}")

        all_words_list.append(best_words)
        all_full_texts.append(best_text)

    if progress_cb: progress_cb(0.65)
    if status_cb: status_cb("🔍 ตรวจจับธนาคาร + Parse รายการ...")

    combined_text = "\n".join(all_full_texts)
    active_bank   = selected_bank if selected_bank != "AUTO" else detect_bank(combined_text)

    words_df = pd.concat(all_words_list, ignore_index=True)

    if progress_cb: progress_cb(0.75)
    parsed_lines = parse_lines(words_df, active_bank)
    dedup_lines  = deduplicate(parsed_lines)
    dedup_lines  = dedup_lines.sort_values(["file_name","y_center"], na_position="last").reset_index(drop=True)

    if progress_cb: progress_cb(0.9)
    if status_cb: status_cb("✅ คำนวณ Balance Chain...")

    parsed_df, check_df = add_debit_credit_check(dedup_lines, active_bank)

    review_df = check_df[
        check_df["balance_check"].astype(str).str.contains(
            "REVIEW|BROKEN|UNKNOWN|NO_PREV|MISSING|SUGGEST|HEALED", regex=True
        )
    ].copy().reset_index(drop=True)

    if progress_cb: progress_cb(1.0)
    return parsed_df, check_df, review_df, active_bank


# ============================================================
# 11)  Excel Export
# ============================================================
def create_summary_df(parsed_df: pd.DataFrame, check_df: pd.DataFrame, active_bank: str) -> pd.DataFrame:
    db_s = pd.to_numeric(parsed_df.get("debit", pd.Series(dtype=float)), errors="coerce")
    cr_s = pd.to_numeric(parsed_df.get("credit", pd.Series(dtype=float)), errors="coerce")

    return pd.DataFrame([
        ["detected_bank", active_bank],
        ["transaction_rows", int(parsed_df["date"].notna().sum()) if not parsed_df.empty else 0],
        ["debit_total", float(db_s.fillna(0).sum())],
        ["credit_total", float(cr_s.fillna(0).sum())],
        ["net_change", float(cr_s.fillna(0).sum() - db_s.fillna(0).sum())],
        ["ok_checks", int(check_df["balance_check"].astype(str).str.contains("OK").sum()) if not check_df.empty and "balance_check" in check_df.columns else 0],
        ["anomalies_found", int(check_df["balance_check"].astype(str).str.contains("REVIEW|BROKEN|MISSING", regex=True).sum()) if not check_df.empty and "balance_check" in check_df.columns else 0],
    ], columns=["metric", "value"])


def prepare_check_export_df(check_df: pd.DataFrame) -> pd.DataFrame:
    """
    ใช้โครงสร้างคอลัมน์ check แบบโค้ดที่ 1:
    - ซ่อน field ที่รก/ใช้ debug OCR ภายใน
    - เปลี่ยนชื่อ expected_balance/diff เป็น expected_balance_py/diff_py
    - เพิ่ม expected_balance_excel/anomaly_reason_excel สำหรับสูตรใน Excel
    """
    export_cols = [
        "seq", "date", "time", "code", "debit", "credit", "balance",
        "prev_balance", "suggested_amount", "suggested_type",
        "expected_balance_py", "diff_py", "order_method",
        "expected_balance_excel", "anomaly_reason_excel",
    ]

    if check_df is None or check_df.empty:
        return pd.DataFrame(columns=export_cols)

    df = check_df.copy()
    df = df.drop(
        columns=[
            "amount", "line_confidence", "min_confidence",
            "money_tokens", "raw_line_text", "balance_check",
        ],
        errors="ignore",
    )
    df = df.rename(columns={"expected_balance": "expected_balance_py", "diff": "diff_py"})
    df["expected_balance_excel"] = None
    df["anomaly_reason_excel"] = None

    for col in export_cols:
        if col not in df.columns:
            df[col] = None

    extra_cols = [c for c in df.columns if c not in export_cols]
    return df[export_cols + extra_cols]


def create_excel(parsed_df: pd.DataFrame, check_df: pd.DataFrame,
                 review_df: pd.DataFrame, active_bank: str) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()

    # บังคับ parsed_stm ให้เป็นคอลัมน์แบบโค้ดที่ 1 เสมอ
    parsed_export_df = parsed_df.copy()
    for col in ["date", "debit", "credit", "balance"]:
        if col not in parsed_export_df.columns:
            parsed_export_df[col] = None
    parsed_export_df = parsed_export_df[["date", "debit", "credit", "balance"]]

    check_export_df = prepare_check_export_df(check_df)
    summary_df = create_summary_df(parsed_export_df, check_df, active_bank)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        parsed_export_df.to_excel(writer, sheet_name="parsed_stm", index=False)
        check_export_df.to_excel(writer,  sheet_name="check",      index=False)
        summary_df.to_excel(writer,      sheet_name="summary",    index=False)
        review_df.to_excel(writer,       sheet_name="ocr_review", index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    # ---------- parsed_stm format ----------
    if "parsed_stm" in wb.sheetnames:
        ws = wb["parsed_stm"]
        headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(color="000000", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for col in ["debit", "credit", "balance"]:
            if col in headers:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=headers[col]).number_format = "#,##0.00"
        ws.freeze_panes = "A2"
        for col_idx in range(1, ws.max_column + 1):
            cl = get_column_letter(col_idx)
            max_len = max((len(str(c.value or "")) for c in ws[cl]), default=10)
            ws.column_dimensions[cl].width = min(max(max_len + 2, 12), 35)

    # ---------- check format + Excel formulas แบบโค้ดที่ 1 ----------
    if "check" in wb.sheetnames:
        ws = wb["check"]
        headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}

        required = [
            "debit", "credit", "balance", "prev_balance", "suggested_amount",
            "expected_balance_excel", "anomaly_reason_excel",
        ]

        if all(c in headers for c in required):
            debit_col     = get_column_letter(headers["debit"])
            credit_col    = get_column_letter(headers["credit"])
            balance_col   = get_column_letter(headers["balance"])
            prev_col      = get_column_letter(headers["prev_balance"])
            suggested_col = get_column_letter(headers["suggested_amount"])
            expected_col  = get_column_letter(headers["expected_balance_excel"])
            reason_col    = get_column_letter(headers["anomaly_reason_excel"])
            tol = BALANCE_TOLERANCE

            for row in range(2, ws.max_row + 1):
                ws[f"{expected_col}{row}"] = (
                    f'=IF({prev_col}{row}="","",ROUND({prev_col}{row}'
                    f'-IF({debit_col}{row}="",0,{debit_col}{row})'
                    f'+IF({credit_col}{row}="",0,{credit_col}{row}),2))'
                )
                ws[f"{reason_col}{row}"] = (
                    f'=IF({prev_col}{row}="","OPENING_OR_NO_PREV",'
                    f'IF({balance_col}{row}="","MISSING_BALANCE",'
                    f'IF({suggested_col}{row}<>"","กระทบยอดให้ตรวจสอบ",'
                    f'IF(AND({debit_col}{row}="",{credit_col}{row}=""),"MISSING_DEBIT_CREDIT",'
                    f'IF(ABS({balance_col}{row}-{expected_col}{row})>{tol},"BALANCE_NOT_MATCH","OK")))))'
                )

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(color="000000", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)

        number_cols = [
            "debit", "credit", "balance", "prev_balance", "suggested_amount",
            "expected_balance_py", "diff_py", "expected_balance_excel",
        ]
        for col in number_cols:
            if col in headers:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=headers[col]).number_format = "#,##0.00"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = None

        if ws.max_row >= 2 and "anomaly_reason_excel" in headers:
            reason_col = get_column_letter(headers["anomaly_reason_excel"])
            light_red_fill = PatternFill("solid", fgColor="FCE4E4")
            dark_red_font = Font(color="9C0006")
            ws.conditional_formatting.add(
                f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
                FormulaRule(
                    formula=[f'=AND(${reason_col}2<>"OK",${reason_col}2<>"OPENING_OR_NO_PREV",${reason_col}2<>"")'],
                    fill=light_red_fill,
                    font=dark_red_font,
                ),
            )

        for col_idx in range(1, ws.max_column + 1):
            cl = get_column_letter(col_idx)
            max_len = max((len(str(c.value or "")) for c in ws[cl]), default=10)
            ws.column_dimensions[cl].width = min(max(max_len + 2, 12), 35)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# 12)  Session State Init
# ============================================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "results" not in st.session_state:
    st.session_state.results = None


# ============================================================
# 13)  LOGIN PAGE
# ============================================================
def login_page():
    st.title("Login")
    st.caption("กรุณาเข้าสู่ระบบก่อนใช้งาน STM Image Parser")

    password = st.text_input("Password", type="password", placeholder="กรอกรหัสผ่าน")

    if st.button("เข้าสู่ระบบ", type="primary", use_container_width=True):
        correct_pw = st.secrets.get("APP_PASSWORD", "stm2025")
        if password == correct_pw:
            st.session_state.authenticated = True
            st.success("เข้าสู่ระบบสำเร็จ")
            st.rerun()
        else:
            st.error("รหัสผ่านไม่ถูกต้อง")

    st.stop()


# ============================================================
# 14)  MAIN APP PAGE
# ============================================================
def render_section(step_no: str, title: str, badge: str | None = None):
    badge_html = f'<span class="badge">{badge}</span>' if badge else ""
    st.markdown(f"""
<div class="section-title">
    <span class="step-number">{step_no}</span>
    <span class="step-text">{title}</span>
    {badge_html}
</div>
""", unsafe_allow_html=True)


def format_money_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def main_app():
    st.title("STM Image Parser")
    st.caption(f"Google Cloud Vision OCR · ปีที่ใช้ {FORCE_YEAR}")
    st.divider()

    render_section("1", "เลือกธนาคาร")
    bank_choice = st.radio(
        "ธนาคาร",
        options=list(BANK_LABELS.keys()),
        format_func=lambda x: BANK_LABELS[x],
        index=0,
        horizontal=True,
        label_visibility="collapsed",
    )

    st.divider()

    selected_short = "AUTO" if bank_choice == "AUTO" else bank_choice
    render_section("2", "อัปโหลดภาพ Statement", selected_short)
    uploaded_files = st.file_uploader(
        "เลือกไฟล์ภาพ Statement (อัปโหลดได้หลายไฟล์พร้อมกัน)",
        type=["jpg", "jpeg", "png"],
        accept_multiple_files=True,
        help="รองรับ JPG / PNG · KBANK / KRUNGSRI / BBL / SCB",
    )

    if not uploaded_files:
        if st.button("ออกจากระบบ", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.results = None
            st.rerun()
        st.stop()

    st.caption(f"ไฟล์ที่เลือก: {len(uploaded_files)} ไฟล์  •  ธนาคาร: {BANK_LABELS.get(bank_choice, bank_choice)}")
    st.divider()

    render_section("3", "ประมวลผล")
    col_btn, col_info = st.columns([2, 3])
    with col_btn:
        run_clicked = st.button("เริ่มประมวลผล", type="primary", use_container_width=True)
    with col_info:
        st.markdown('<div class="muted-note"></div>', unsafe_allow_html=True)

    col_clear, col_logout = st.columns([1, 1])
    with col_clear:
        if st.button("ล้างผลลัพธ์", use_container_width=True):
            st.session_state.results = None
            st.rerun()
    with col_logout:
        if st.button("ออกจากระบบ", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.results = None
            st.rerun()

    if run_clicked:
        st.session_state.results = None
        progress_bar = st.progress(0, text="กำลังเริ่มประมวลผล...")
        status_box = st.empty()

        def prog_cb(val):
            progress_bar.progress(val, text=f"{int(val * 100)}%")

        def stat_cb(msg):
            status_box.info(str(msg))

        try:
            parsed_df, check_df, review_df, active_bank = run_pipeline(
                uploaded_files, bank_choice, prog_cb, stat_cb
            )
            excel_bytes = create_excel(parsed_df, check_df, review_df, active_bank)
            excel_name = f"stm_result_{active_bank.lower()}.xlsx"
            st.session_state.results = {
                "parsed": parsed_df,
                "check": check_df,
                "review": review_df,
                "active_bank": active_bank,
                "excel": excel_bytes,
                "excel_name": excel_name,
            }
            progress_bar.progress(1.0, text="ประมวลผลเสร็จสิ้น ✅")
            status_box.success(f"ประมวลผลสำเร็จ พบ {len(parsed_df):,} รายการ · ธนาคาร: {active_bank}")
        except Exception as e:
            progress_bar.empty()
            status_box.error(f"อ่านไฟล์ไม่สำเร็จ: {e}")
            st.exception(e)

    # ─── Results Section ──────────────────────────────────────
    if st.session_state.results:
        res = st.session_state.results
        parsed_df   = res["parsed"]
        check_df    = res["check"]
        review_df   = res["review"]
        active_bank = res["active_bank"]
        excel_bytes = res["excel"]
        excel_name  = res["excel_name"]

        st.divider()
        st.success(f"ประมวลผลสำเร็จ! พบ {len(parsed_df):,} รายการ")
        st.caption(f"ธนาคารที่ตรวจพบ/เลือกใช้: {active_bank}")

        # ── Download button (manual, prominent) ───────────────
        st.download_button(
            label="📥 ดาวน์โหลด Excel",
            data=excel_bytes,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        st.divider()

        db_s  = format_money_series(parsed_df.get("debit",  pd.Series(dtype=float)))
        cr_s  = format_money_series(parsed_df.get("credit", pd.Series(dtype=float)))
        bal_s = pd.to_numeric(parsed_df.get("balance", pd.Series(dtype=float)), errors="coerce")
        net     = cr_s.sum() - db_s.sum()
        last_bal = float(bal_s.dropna().iloc[-1]) if len(bal_s.dropna()) else 0.0
        ok_n    = int(check_df["balance_check"].astype(str).str.contains("OK").sum()) if not check_df.empty else 0
        rv_n    = len(review_df)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("จำนวนรายการ",      f"{len(parsed_df):,}")
        c2.metric("ยอดคงเหลือล่าสุด", f"{last_bal:,.2f}")
        c3.metric("ผ่านตรวจสอบ",      f"{ok_n:,}")
        c4.metric("ต้องตรวจ",         f"{rv_n:,}")

        c5, c6, c7 = st.columns(3)
        c5.metric("รวมเดบิต",  f"{db_s.sum():,.2f}")
        c6.metric("รวมเครดิต", f"{cr_s.sum():,.2f}")
        c7.metric("ยอดสุทธิ",  f"{net:,.2f}")

        st.markdown("#### ตัวอย่างรายการ")
        if parsed_df.empty:
            st.warning("ไม่พบรายการธุรกรรม")
        else:
            preview_df = parsed_df.head(30).copy()
            for col in ["debit", "credit", "balance"]:
                if col in preview_df.columns:
                    preview_df[col] = pd.to_numeric(preview_df[col], errors="coerce")
            st.dataframe(
                preview_df.style.format({
                    "debit":   lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                    "credit":  lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                    "balance": lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                }),
                use_container_width=True,
                height=400,
            )

        st.divider()
        st.markdown("#### รายละเอียดผลลัพธ์")
        tab1, tab2, tab3 = st.tabs(["parsed_stm", "check", "OCR Review"])

        with tab1:
            # แสดงคอลัมน์แบบโค้ดที่ 1: date / debit / credit / balance
            if parsed_df.empty:
                st.warning("ไม่พบรายการธุรกรรม")
            else:
                display_df = parsed_df[["date", "debit", "credit", "balance"]].copy()
                for col in ["debit", "credit", "balance"]:
                    display_df[col] = pd.to_numeric(display_df[col], errors="coerce")

                st.dataframe(
                    display_df.style.format({
                        "debit":   lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                        "credit":  lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                        "balance": lambda v: f"{v:,.2f}" if pd.notna(v) else "",
                    }),
                    use_container_width=True,
                    height=520,
                )

        with tab2:
            # แสดง check ตามคอลัมน์ export แบบโค้ดที่ 1
            check_display_df = prepare_check_export_df(check_df)
            if check_display_df.empty:
                st.info("ไม่มีข้อมูล")
            else:
                show_cols = [
                    "seq", "date", "time", "code", "debit", "credit", "balance",
                    "prev_balance", "suggested_amount", "suggested_type",
                    "expected_balance_py", "diff_py", "order_method",
                    "expected_balance_excel", "anomaly_reason_excel",
                ]
                show_cols = [c for c in show_cols if c in check_display_df.columns]
                st.dataframe(check_display_df[show_cols], use_container_width=True, height=520)

        with tab3:
            # เก็บ raw_line_text / balance_check ไว้ตรวจ OCR ตามเดิม
            if review_df.empty:
                st.success("ไม่มีรายการที่ต้องตรวจสอบ")
            else:
                st.warning(f"พบ {len(review_df):,} รายการที่ต้องตรวจสอบ")
                show_cols = ["seq", "date", "debit", "credit", "balance", "balance_check", "raw_line_text"]
                show_cols = [c for c in show_cols if c in review_df.columns]
                st.dataframe(review_df[show_cols], use_container_width=True, height=420)

        st.divider()

        # ── Bottom download button (second copy for convenience) ──
        st.download_button(
            label="📥 ดาวน์โหลด Excel",
            data=excel_bytes,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_bottom",
            use_container_width=True,
        )

        if st.button("ล้างข้อมูลหลังใช้งาน", use_container_width=True):
            st.session_state.results = None
            st.success("ล้างข้อมูล Session แล้ว")
            st.rerun()

    st.divider()
    st.caption("รองรับ: KBANK · KRUNGSRI · BBL · SCB")


# ============================================================
# 15)  Entry Point
# ============================================================
if not st.session_state.authenticated:
    login_page()
else:
    main_app()
